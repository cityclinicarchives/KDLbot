
import base64
import json
import mimetypes
import os
import re
import tempfile
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any

from openai import OpenAI
from rapidfuzz import fuzz
from openpyxl import load_workbook

from config import (
    OPENAI_API_KEY,
    OPENAI_TEXT_MODEL,
    OPENAI_VISION_MODEL,
    OPENAI_AUDIO_MODEL,
)
from pricing import (
    LabTest,
    MatchedItem,
    normalize_text,
    split_user_tests,
    match_tests,
    find_candidate_tests,
    labtest_to_candidate,
    make_matched_from_code,
    make_included_item,
)


MEMORY_PATH = Path("cleaner_memory.json")


@dataclass
class AIUsage:
    used: bool = False
    source: str = ""
    prompt_tokens: int = 0
    completion_tokens: int = 0
    total_tokens: int = 0

    def add(self, other: "AIUsage") -> None:
        if not other:
            return
        self.used = self.used or other.used
        sources = [s for s in [self.source, other.source] if s]
        self.source = "+".join(dict.fromkeys(sources))
        self.prompt_tokens += int(other.prompt_tokens or 0)
        self.completion_tokens += int(other.completion_tokens or 0)
        self.total_tokens += int(other.total_tokens or 0)

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class ExtractedTextResult:
    raw_text: str
    usage: AIUsage


@dataclass
class AIMatchResult:
    matched_items: list[MatchedItem]
    cleaned_items: list[str]
    raw_text: str
    usage: AIUsage
    used_gpt_matching: bool = False


def is_ai_enabled() -> bool:
    return bool(OPENAI_API_KEY)


def _client() -> OpenAI:
    if not OPENAI_API_KEY:
        raise RuntimeError("OPENAI_API_KEY не задан")
    return OpenAI(api_key=OPENAI_API_KEY)


def _usage_from_chat_response(response, source: str) -> AIUsage:
    usage = getattr(response, "usage", None)
    if not usage:
        return AIUsage(used=True, source=source)
    prompt = int(getattr(usage, "prompt_tokens", 0) or 0)
    completion = int(getattr(usage, "completion_tokens", 0) or 0)
    total = int(getattr(usage, "total_tokens", 0) or (prompt + completion))
    return AIUsage(
        used=True,
        source=source,
        prompt_tokens=prompt,
        completion_tokens=completion,
        total_tokens=total,
    )


def _usage_from_audio_response(response, source: str) -> AIUsage:
    usage = getattr(response, "usage", None)
    if not usage:
        return AIUsage(used=True, source=source)
    # В разных версиях SDK usage может быть объектом с разными именами полей.
    prompt = int(getattr(usage, "input_tokens", 0) or getattr(usage, "prompt_tokens", 0) or 0)
    completion = int(getattr(usage, "output_tokens", 0) or getattr(usage, "completion_tokens", 0) or 0)
    total = int(getattr(usage, "total_tokens", 0) or (prompt + completion))
    return AIUsage(True, source, prompt, completion, total)


def _load_memory() -> dict:
    if not MEMORY_PATH.exists():
        return {"cleaned": {}, "choices": {}}
    try:
        data = json.loads(MEMORY_PATH.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data.setdefault("cleaned", {})
            data.setdefault("choices", {})
            return data
    except Exception:
        pass
    return {"cleaned": {}, "choices": {}}


def _save_memory(data: dict) -> None:
    try:
        MEMORY_PATH.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        # На некоторых хостингах файловая система может быть временной или read-only.
        pass


def remember_manual_choice(input_text: str, selected_code: str, selected_name: str) -> None:
    key = normalize_text(input_text)
    if not key or not selected_code:
        return
    data = _load_memory()
    data["choices"][key] = {"code": selected_code, "name": selected_name}
    _save_memory(data)


FILLER_PHRASES = [
    "ну", "так", "значит", "пожалуйста", "пожалуй", "наверное", "скорее всего",
    "я хочу", "хочу", "мне нужно", "мне надо", "надо", "нужно", "сдать",
    "сделать", "проверить", "посмотреть", "анализы", "анализ", "еще", "ещё",
    "и еще", "а еще", "также", "плюс", "можно", "давайте",
]


SEARCH_EQUIVALENTS = {
    "орешник": ["лещина", "лещина обыкновенная", "лесной орех", "фундук", "corylus", "corylus avellana"],
    "лещина": ["орешник", "лещина обыкновенная", "лесной орех", "фундук", "corylus", "corylus avellana"],
    # Важно: Candida НЕ является бытовым синонимом "плесень" в запросах на аллергены.
    "плесень": ["плесневый гриб", "плесневые грибы", "alternaria", "alternaria alternata", "cladosporium", "aspergillus", "penicillium"],
    "береза": ["береза", "береза бородавчатая", "betula", "betula alba"],
    "ольха": ["ольха", "ольха серая", "alnus", "alnus incana"],
    "капрограмма": ["копрограмма", "общий анализ кала"],
    "копрограмма": ["копрограмма", "общий анализ кала"],
    "общий белок": ["общий белок"],
    "глюкоза": ["глюкоза"],
    "пт": ["протромбин", "протромбиновое"],
    "тв": ["тромбиновое время"],
    "фг": ["фибриноген"],
    "ковид": ["covid", "covid-19", "коронавирус", "sars-cov-2", "sars cov 2", "мазок на ковид", "пцр на ковид"],
    "covid": ["ковид", "covid-19", "коронавирус", "sars-cov-2", "sars cov 2", "мазок на ковид", "пцр на ковид"],
    "covid-19": ["ковид", "covid", "коронавирус", "sars-cov-2", "sars cov 2", "мазок на ковид", "пцр на ковид"],
    "коронавирус": ["ковид", "covid", "covid-19", "sars-cov-2", "sars cov 2", "мазок на ковид", "пцр на ковид"],
    "sars-cov-2": ["ковид", "covid", "covid-19", "коронавирус", "sars cov 2", "мазок на ковид", "пцр на ковид"],
    "sars cov 2": ["ковид", "covid", "covid-19", "коронавирус", "sars-cov-2", "мазок на ковид", "пцр на ковид"],
}


# Универсальные смысловые эквиваленты для candidate search.
# Это НЕ замена GPT, а способ дать GPT правильный набор вариантов из всего прайса.
UNIVERSAL_SEARCH_EQUIVALENTS = {
    "базовая биохимия": [
        "биохимический анализ крови базовый",
        "биохимический анализ крови, базовый",
        "базовый биохимический анализ крови",
        "биохимия базовая",
    ],
    "биохимия базовая": [
        "биохимический анализ крови базовый",
        "биохимический анализ крови, базовый",
        "базовая биохимия",
    ],
    "биохимия": [
        "биохимический анализ крови",
        "биохимический анализ крови базовый",
        "биохимический анализ крови, базовый",
    ],
    "пцр на 12 инфекций": [
        "пцр-12",
        "пцр 12",
        "пцр панель 12",
        "пцр-диагностика 12",
        "урогенитальные инфекции 12",
    ],
    "пцр 12 инфекций": [
        "пцр-12",
        "пцр 12",
        "пцр панель 12",
        "урогенитальные инфекции 12",
    ],
    "12 инфекций": [
        "пцр-12",
        "пцр 12",
        "урогенитальные инфекции 12",
    ],
    "впч высокого онкогенного риска": [
        "впч высокого онкогенного риска",
        "папилломавирусов высокого онкогенного риска",
        "днк папилломавирусов",
    ],
    "высокого онкогенного риска": [
        "впч высокого онкогенного риска",
        "папилломавирусов высокого онкогенного риска",
    ],
    "мазок на ковид": [
        "covid",
        "covid-19",
        "ковид",
        "коронавирус",
        "sars-cov-2",
        "sars cov 2",
        "днк sars-cov-2",
        "рнк sars-cov-2",
        "пцр covid",
        "пцр на covid",
        "пцр на ковид",
    ],
    "пцр на ковид": [
        "covid",
        "covid-19",
        "ковид",
        "коронавирус",
        "sars-cov-2",
        "sars cov 2",
        "днк sars-cov-2",
        "рнк sars-cov-2",
    ],
    "ковид": [
        "covid",
        "covid-19",
        "коронавирус",
        "sars-cov-2",
        "sars cov 2",
        "днк sars-cov-2",
        "рнк sars-cov-2",
    ],
}

PHRASE_TO_CANONICAL = {
    "общий анализ крови": "ОАК",
    "клинический анализ крови": "ОАК",
    "анализ крови": "ОАК",
    "оак": "ОАК",
    "общий анализ мочи": "ОАМ",
    "анализ мочи": "ОАМ",
    "оам": "ОАМ",
    "моча": "ОАМ",
    "капрограмма": "копрограмма",
    "копрограмма": "копрограмма",
    "общий анализ кала": "копрограмма",
    "алт": "АЛТ",
    "аст": "АСТ",
    "ттг": "ТТГ",
    "мно": "МНО",
    "ачтв": "АЧТВ",
    "тв": "ТВ",
    "фг": "ФГ",
    "фибриноген": "Фибриноген",
    "д димер": "Д-димер",
    "д-димер": "Д-димер",
    "коагулограмма": "Коагулограмма",
    "глюкоза": "Глюкоза",
    "общий белок": "Общий белок",
    "базовая биохимия": "базовая биохимия",
    "биохимия базовая": "базовая биохимия",
    "пцр на 12 инфекций": "ПЦР на 12 инфекций",
    "пцр 12 инфекций": "ПЦР на 12 инфекций",
    "12 инфекций": "ПЦР на 12 инфекций",
    "впч высокого онкогенного риска": "ВПЧ высокого онкогенного риска",
    "мазок на ковид": "мазок на ковид",
    "пцр на ковид": "мазок на ковид",
    "ковид": "мазок на ковид",
    "covid": "мазок на ковид",
    "covid-19": "мазок на ковид",
    "коронавирус": "мазок на ковид",
    "sars-cov-2": "мазок на ковид",
    "sars cov 2": "мазок на ковид",
    "береза": "аллерген береза",
    "берёза": "аллерген береза",
    "ольха": "аллерген ольха",
    "орешник": "аллерген лещина",
    "лещина": "аллерген лещина",
    "плесень": "аллерген плесневый гриб",
}

GENERIC_SHORTLIST_TOKENS = {
    "общий", "общая", "анализ", "анализы", "крови", "кровь", "мочи", "моча", "аллерген", "аллергены",
    "сдать", "посчитать", "пожалуйста", "мне", "нужно", "надо", "хочу", "еще", "ещё"
}

ALLERGEN_QUERY_WORDS = {
    "береза", "берёза", "ольха", "орешник", "лещина", "плесень", "плесневый", "плесневые"
}

ALLERGEN_PANEL_WORDS = ["панель", "аллергокомплекс", "комплекс"]


def clean_recognized_text_locally(text: str) -> list[str]:
    raw = str(text or "").strip()
    if not raw:
        return []

    data = _load_memory()
    memory_key = normalize_text(raw)
    if memory_key in data.get("cleaned", {}):
        stored = data["cleaned"][memory_key]
        if isinstance(stored, list) and stored:
            return [str(x).strip() for x in stored if str(x).strip()]

    norm = normalize_text(raw)

    extracted: list[str] = []
    # Сначала вытаскиваем устойчивые медицинские фразы.
    for phrase, canonical in sorted(PHRASE_TO_CANONICAL.items(), key=lambda pair: len(pair[0]), reverse=True):
        if re.search(rf"(^|\s){re.escape(normalize_text(phrase))}($|\s)", norm):
            if canonical not in extracted:
                extracted.append(canonical)

    # Потом делим остаток на части и удаляем слова-паразиты.
    separators = r"\b(?:и еще|и ещё|а еще|а ещё|также|плюс|,|;|\n|\.|\+)\b"
    parts = re.split(separators, raw, flags=re.IGNORECASE)

    for part in parts:
        value = normalize_text(part)
        for filler in sorted(FILLER_PHRASES, key=len, reverse=True):
            value = re.sub(rf"(^|\s){re.escape(normalize_text(filler))}($|\s)", " ", value)
        value = re.sub(r"\s+", " ", value).strip(" -—,.;:")

        if not value or len(value) < 2:
            continue

        canonical = PHRASE_TO_CANONICAL.get(value, value)
        if canonical not in extracted:
            extracted.append(canonical)

    # Не возвращаем слишком длинные разговорные куски как отдельный анализ.
    cleaned = []
    for item in extracted:
        tokens = normalize_text(item).split()
        if len(tokens) > 8:
            continue
        if item not in cleaned:
            cleaned.append(item)

    return cleaned


def remember_cleaned_text(raw_text: str, cleaned_items: list[str]) -> None:
    key = normalize_text(raw_text)
    if not key or not cleaned_items:
        return
    data = _load_memory()
    data["cleaned"][key] = cleaned_items[:20]
    _save_memory(data)


def _resize_image_for_vision(path: str) -> str:
    try:
        from PIL import Image
    except Exception:
        return path

    try:
        image = Image.open(path)
        image = image.convert("RGB")
        max_side = 1280
        w, h = image.size
        scale = min(max_side / max(w, h), 1.0)
        if scale < 1.0:
            image = image.resize((int(w * scale), int(h * scale)))

        fd, out_path = tempfile.mkstemp(suffix=".jpg")
        os.close(fd)
        image.save(out_path, "JPEG", quality=75, optimize=True)
        return out_path
    except Exception:
        return path


def _encode_image_as_data_url(path: str) -> str:
    mime = mimetypes.guess_type(path)[0] or "image/jpeg"
    with open(path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode("utf-8")
    return f"data:{mime};base64,{encoded}"


async def extract_tests_from_audio_file(path: str) -> ExtractedTextResult:
    client = _client()
    with open(path, "rb") as f:
        response = client.audio.transcriptions.create(
            model=OPENAI_AUDIO_MODEL,
            file=f,
        )
    raw_text = str(getattr(response, "text", "") or "").strip()
    return ExtractedTextResult(
        raw_text=raw_text,
        usage=_usage_from_audio_response(response, "audio_transcription"),
    )


async def extract_tests_from_image_file(path: str) -> ExtractedTextResult:
    client = _client()
    resized_path = _resize_image_for_vision(path)
    data_url = _encode_image_as_data_url(resized_path)

    prompt = (
        "Извлеки с изображения только текст со списком медицинских анализов. "
        "Не сопоставляй с прайсом. Не добавляй пояснения. "
        "Верни простой текст, каждый анализ или строку назначения с новой строки."
    )

    response = client.chat.completions.create(
        model=OPENAI_VISION_MODEL,
        temperature=0,
        max_tokens=400,
        messages=[
            {"role": "system", "content": "Ты OCR-модуль для медицинских назначений."},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            },
        ],
    )
    raw_text = response.choices[0].message.content.strip()
    return ExtractedTextResult(
        raw_text=raw_text,
        usage=_usage_from_chat_response(response, "vision_ocr"),
    )



def _phrase_variants_for_search(query: str) -> list[str]:
    """
    Расширяет пользовательский пункт смысловыми вариантами.
    Например:
    "базовая биохимия" -> "биохимический анализ крови, базовый"
    "ПЦР на 12 инфекций" -> "ПЦР-12"
    """
    variants = [query]
    qn = normalize_text(query)

    for key, values in SEARCH_EQUIVALENTS.items():
        if normalize_text(key) in qn:
            variants.extend(values)

    for key, values in UNIVERSAL_SEARCH_EQUIVALENTS.items():
        if normalize_text(key) in qn:
            variants.extend(values)

    # Числовые паттерны: "ПЦР на 12 инфекций" -> "ПЦР-12"
    pcr_num = re.search(r"\bпцр\b.*?\b(\d{1,2})\b", qn)
    if pcr_num:
        n = pcr_num.group(1)
        variants.extend([f"ПЦР-{n}", f"ПЦР {n}", f"ПЦР-{n},", f"{n} инфекций"])

    # "ВПЧ высокого онкогенного риска" часто называется через ДНК папилломавирусов.
    if "впч" in qn and ("онког" in qn or "канцер" in qn):
        variants.extend([
            "ДНК папилломавирусов высокого онкогенного риска",
            "ВПЧ высокого онкогенного риска",
            "папилломавирусов высокого онкогенного риска",
        ])

    deduped = []
    seen = set()
    for v in variants:
        key = normalize_text(v)
        if key and key not in seen:
            seen.add(key)
            deduped.append(v)
    return deduped


def _read_complex_rows_from_xlsx(path: str = "complex_price_list.xlsx") -> list[LabTest]:
    """
    Читает complex_price_list.xlsx напрямую, даже если pricing.load_price_list()
    по какой-то причине не добавил комплексы в общий price_list.
    """
    file_path = Path(path)
    if not file_path.exists():
        return []

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(c or "").strip() for c in rows[0]]

        def get(row, name, default=""):
            if name not in headers:
                return default
            i = headers.index(name)
            if i >= len(row):
                return default
            return default if row[i] is None else row[i]

        result: list[LabTest] = []
        for row in rows[1:]:
            code = str(get(row, "code", "")).strip()
            name = str(get(row, "name", "")).strip()
            if not code and not name:
                continue

            result.append(
                LabTest(
                    code=code,
                    name=name,
                    result_type=str(get(row, "result_type", "комплекс")).strip() or "комплекс",
                    max_discount=0,
                    term_days=str(get(row, "term_days", "")).strip(),
                    price=int(float(str(get(row, "price", 0) or 0).replace(",", "."))),
                    synonyms=[],
                    short_name=str(get(row, "short_name", "")).strip(),
                    group=str(get(row, "group", "")).strip(),
                    priority=int(float(str(get(row, "priority", 0) or 0).replace(",", "."))),
                    auto_select_rule=str(get(row, "auto_select_rule", "")).strip(),
                    components=str(get(row, "components", "")).strip(),
                )
            )
        return result
    except Exception:
        return []


def _semantic_score(query: str, item: LabTest, *, complex_mode: bool = False) -> int:
    """
    Универсальная оценка кандидата.
    Смотрим name, short_name, code, auto_select_rule и components комплекса.
    """
    variants = _phrase_variants_for_search(query)
    best = 0

    item_texts = [
        item.name,
        getattr(item, "short_name", ""),
        getattr(item, "code", ""),
        getattr(item, "auto_select_rule", ""),
    ]
    if complex_mode and getattr(item, "components", ""):
        item_texts.append(item.components)

    for q in variants:
        qn = normalize_text(q)
        if not qn:
            continue
        for item_text in item_texts:
            tn = normalize_text(item_text)
            if not tn:
                continue
            score = fuzz.WRatio(qn, tn)
            score = max(score, fuzz.token_set_ratio(qn, tn))
            if qn in tn:
                score = max(score, 96)
            # "пцр 12" и "пцр-12" после normalize_text становятся одинаковыми.
            best = max(best, int(score))

    # Бонус за общий смысл: "базовая биохимия" <-> "биохимический анализ крови, базовый"
    qn_all = normalize_text(" ".join(variants))
    name_n = normalize_text(item.name)
    comp_n = normalize_text(getattr(item, "components", ""))

    if "биохим" in qn_all and "биохим" in name_n:
        best = max(best, 88)
        if "базов" in qn_all and "базов" in name_n:
            best = max(best, 115)

    if "пцр" in qn_all and "пцр" in name_n:
        best = max(best, 88)
        pcr_num = re.search(r"\bпцр\b.*?\b(\d{1,2})\b", qn_all)
        if pcr_num and re.search(rf"\bпцр\s*{pcr_num.group(1)}\b", name_n):
            best = max(best, 120)

    if "впч" in qn_all and ("впч" in name_n or "папилломавирус" in name_n):
        best = max(best, 90)
        if "онког" in qn_all and ("онког" in name_n or "канцер" in name_n):
            best = max(best, 115)

    if _is_covid_query_text(qn_all):
        covid_name_hit = any(token in name_n for token in [
            "ковид", "covid", "коронавирус", "sars cov", "sars cov 2", "sars"
        ])
        covid_comp_hit = any(token in comp_n for token in [
            "ковид", "covid", "коронавирус", "sars cov", "sars cov 2", "sars"
        ])
        if covid_name_hit or covid_comp_hit:
            best = max(best, 120)
            # Предпочитаем ПЦР/мазок/РНК-ДНК SARS-CoV-2, если это видно в названии.
            if any(token in name_n for token in ["пцр", "мазок", "рнк", "днк"]):
                best = max(best, 130)

    # Для комплекса учитываем состав.
    if complex_mode and comp_n:
        q_tokens = {t for t in qn_all.split() if len(t) >= 3 and t not in GENERIC_SHORTLIST_TOKENS}
        comp_tokens = set(comp_n.split())
        shared = len(q_tokens & comp_tokens)
        if shared:
            best = max(best, min(100, shared * 12))

    return int(best)


def _dedupe_labtests(items: list[tuple[int, LabTest]], limit: int) -> list[tuple[int, LabTest]]:
    seen = set()
    result = []
    for score, item in sorted(items, key=lambda pair: (-pair[0], -pair[1].priority, pair[1].price, pair[1].name)):
        key = item.code or normalize_text(item.name)
        if key in seen:
            continue
        seen.add(key)
        result.append((score, item))
        if len(result) >= limit:
            break
    return result



def _catalog_item(item: LabTest) -> dict[str, Any]:
    return {
        "code": item.code,
        "name": item.name,
        "price": item.price,
        "term_days": item.term_days,
        "result_type": item.result_type,
        "priority": item.priority,
        "components": item.components if item.result_type == "комплекс" else "",
    }


def _candidate_score(query: str, item: LabTest, *, complex_mode: bool = False) -> int:
    q = normalize_text(query)
    name = normalize_text(item.name)
    score = fuzz.WRatio(q, name)
    if q and q in name:
        score = max(score, 92)

    # Для комплексов состав используем только для попадания в shortlist и дедубликации.
    if complex_mode and item.components:
        comp = normalize_text(item.components)
        comp_score = fuzz.token_set_ratio(q, comp)
        shared = len(set(q.split()) & set(comp.split()))
        score = max(score, int(comp_score), min(100, shared * 12))
    return int(score)


def _is_likely_allergen_query(text: str) -> bool:
    qn = normalize_text(text)
    return any(word in qn for word in ALLERGEN_QUERY_WORDS)


def _is_panel_or_complex_allergen(item: LabTest) -> bool:
    name = normalize_text(item.name)
    return any(word in name for word in ALLERGEN_PANEL_WORDS)


def _has_meaningful_token_overlap(query_text: str, item_name: str) -> bool:
    q_tokens = {t for t in normalize_text(query_text).split() if len(t) >= 4 and t not in GENERIC_SHORTLIST_TOKENS}
    if not q_tokens:
        return False
    item_norm = normalize_text(item_name)
    return any(tok in item_norm for tok in q_tokens)


def _force_item_by_name_contains(price_list: list[LabTest], scored: dict[str, tuple[int, LabTest]], words: list[str], *, score: int = 100, exclude_panels: bool = False, prefer_cheaper: bool = True) -> None:
    matches = []
    for item in price_list:
        name_norm = normalize_text(item.name)
        if all(w in name_norm for w in words):
            if exclude_panels and _is_panel_or_complex_allergen(item):
                continue
            matches.append(item)
    if not matches:
        return
    # Для аллергенов обычно в прайсе есть 2 версии: обычная ИФА дешевле и ImmunoCAP дороже.
    # По умолчанию пациент ожидает базовый индивидуальный аллерген, поэтому ставим более дешевый выше.
    matches.sort(key=lambda x: (x.price if prefer_cheaper else -x.price, x.name))
    for item in matches[:5]:
        scored[item.code] = (score, item)


def _build_candidate_catalog(raw_text: str, cleaned_items: list[str], price_list: list[LabTest]) -> dict[str, list[dict[str, Any]]]:
    """
    Универсальный candidate search по ВСЕМУ price_list + complex_price_list.

    Почему это важно:
    GPT nano не должен угадывать из воздуха. Он должен выбирать из релевантного набора.
    Поэтому здесь Python:
    1) разбивает запрос на пункты;
    2) расширяет каждый пункт смысловыми вариантами;
    3) ищет кандидатов по всему прайсу и по всем комплексам;
    4) передает GPT только сильные варианты.
    """
    base_queries = []
    base_queries.extend(split_user_tests(raw_text))
    base_queries.extend(cleaned_items)
    if not base_queries:
        base_queries = [raw_text]

    # Не теряем исходный текст целиком: он нужен для понимания комплексов.
    full_query_text = " ".join([raw_text] + cleaned_items)
    qn_full = normalize_text(full_query_text)

    # Подмешиваем комплексы из файла напрямую, если их нет в общем price_list.
    complex_from_file = _read_complex_rows_from_xlsx()
    existing_codes = {item.code for item in price_list if item.code}
    full_price_list = list(price_list)
    for item in complex_from_file:
        if item.code and item.code not in existing_codes:
            full_price_list.append(item)
            existing_codes.add(item.code)

    individual_items = [item for item in full_price_list if item.result_type != "комплекс"]
    complex_items = [item for item in full_price_list if item.result_type == "комплекс"]

    individual_scored: list[tuple[int, LabTest]] = []
    complex_scored: list[tuple[int, LabTest]] = []

    # --- 1. Универсальный поиск по каждому пункту ---
    for query in base_queries:
        if not str(query).strip():
            continue

        for item in individual_items:
            # Если запрос про отдельные аллергены, не даем панелям забивать shortlist.
            if _is_likely_allergen_query(query) and _is_panel_or_complex_allergen(item):
                continue
            score = _semantic_score(query, item, complex_mode=False)
            if score >= 62:
                individual_scored.append((score, item))

        for item in complex_items:
            # Если запрос про отдельные аллергены, не добавляем аллергопанели, если их прямо не просили.
            name_norm = normalize_text(item.name)
            if _is_likely_allergen_query(query) and any(w in name_norm for w in ["панель", "аллергокомплекс", "комплекс"]):
                if not any(x in normalize_text(query) for x in ["панель", "комплекс"]):
                    continue
            score = _semantic_score(query, item, complex_mode=True)
            if score >= 60:
                complex_scored.append((score, item))

    # --- 2. Жесткие безопасные дефолты, где медицинский смысл важнее fuzzy ---
    if "оак" in qn_full or "общий анализ крови" in qn_full or "клинический анализ крови" in qn_full or "анализ крови" in qn_full:
        for item in individual_items:
            n = normalize_text(item.name)
            if "клинический анализ крови" in n and "соэ" in n and "веноз" in n:
                individual_scored.append((130, item))
            elif "клинический анализ крови" in n and "соэ" in n:
                individual_scored.append((120, item))

    if "оам" in qn_full or "общий анализ мочи" in qn_full:
        for item in individual_items:
            n = normalize_text(item.name)
            if "общий анализ мочи" in n:
                individual_scored.append((125, item))

    if "копрограмм" in qn_full or "капрограмм" in qn_full or "общий анализ кала" in qn_full:
        for item in individual_items:
            n = normalize_text(item.name)
            if "копрограмма" in n:
                individual_scored.append((125, item))

    if "глюкоза" in qn_full:
        for item in individual_items:
            n = normalize_text(item.name)
            if n == "глюкоза" or n.startswith("глюкоза "):
                individual_scored.append((120, item))

    if "общий белок" in qn_full:
        for item in individual_items:
            n = normalize_text(item.name)
            if "общий белок" in n:
                individual_scored.append((120, item))

    # Базовая биохимия: ищем комплекс по всему complex_price_list.
    if "биохим" in qn_full:
        for item in complex_items:
            n = normalize_text(item.name)
            if "биохимический анализ крови" in n and "базов" in n:
                complex_scored.append((135, item))
            elif "биохим" in n and "базов" in n:
                complex_scored.append((125, item))
            elif "биохимический анализ крови" == n:
                complex_scored.append((95, item))

    # ПЦР на N инфекций: ПЦР-12, ПЦР-15 и т.д.
    pcr_num = re.search(r"\bпцр\b.*?\b(\d{1,2})\b", qn_full)
    if pcr_num:
        n_requested = pcr_num.group(1)
        for item in complex_items:
            n = normalize_text(item.name)
            if re.search(rf"\bпцр\s*{re.escape(n_requested)}\b", n):
                # Неспецифические варианты выше, специфические биоматериалы ниже.
                penalty = 0
                if any(x in n for x in ["эякулят", "моча"]):
                    penalty = 10
                complex_scored.append((140 - penalty, item))

    # Коагулограмма и компоненты.
    coag_markers = ["коагул", "мно", "ачтв", "протромбин", "пт", "тв", "тромбиновое", "фг", "фибриноген"]
    if any(m in qn_full for m in coag_markers):
        for item in complex_items:
            n = normalize_text(item.name)
            if "коагулограмма" in n and "скрининг" in n:
                complex_scored.append((140, item))
            elif "коагулограмма" in n:
                complex_scored.append((100, item))
        for item in individual_items:
            n = normalize_text(item.name)
            if any(m in n for m in ["мно", "ачтв", "протромбин", "тромбиновое", "фибриноген"]):
                individual_scored.append((95, item))

    # COVID / SARS-CoV-2.
    if _is_covid_query_text(qn_full):
        for item in individual_items:
            n = normalize_text(item.name)
            blob = normalize_text(item.name + " " + getattr(item, "short_name", "") + " " + getattr(item, "auto_select_rule", ""))
            if any(token in blob for token in ["ковид", "covid", "коронавирус", "sars cov", "sars cov 2", "sars"]):
                score = 125
                if any(token in blob for token in ["пцр", "мазок", "рнк", "днк"]):
                    score = 135
                individual_scored.append((score, item))
        for item in complex_items:
            blob = normalize_text(item.name + " " + getattr(item, "components", ""))
            if any(token in blob for token in ["ковид", "covid", "коронавирус", "sars cov", "sars cov 2", "sars"]):
                complex_scored.append((115, item))

    # Аллергены.
    allergen_map = {
        "берез": ["береза", "betula"],
        "ольх": ["ольха", "alnus"],
        "ореш": ["лещина", "corylus"],
        "лещин": ["лещина", "corylus"],
        "плес": ["плеснев", "alternaria"],
    }
    for trigger, needles in allergen_map.items():
        if trigger in qn_full:
            for item in individual_items:
                n = normalize_text(item.name)
                if _is_panel_or_complex_allergen(item):
                    continue
                if any(needle in n for needle in needles):
                    score = 125
                    if trigger == "плес" and "alternaria" in n:
                        score = 135
                    individual_scored.append((score, item))

    individuals = _dedupe_labtests(individual_scored, limit=120)
    complexes = _dedupe_labtests(complex_scored, limit=80)

    return {
        "individual_analyses": [_catalog_item(item) for _, item in individuals],
        "complexes": [_catalog_item(item) for _, item in complexes],
    }


GPT_MATCHING_SYSTEM_PROMPT = """
Ты — медицинский эксперт лабораторного прайс-листа. Твоя задача — сопоставить запрос пациента с анализами и комплексами из переданного каталога.

Строгие правила:
1. Используй только переданные позиции каталога. Не придумывай анализы.
2. Для поиска ориентируйся на поле name. Не используй synonyms, их нет в каталоге.
3. Для комплексов используй поле components, чтобы понять, какие отдельные показатели входят в комплекс.
4. Если пользователь просит набор показателей, который очевидно входит в комплекс, выбери комплекс.
5. Если пользователь одновременно просит комплекс и отдельный анализ, который входит в этот комплекс, НЕ удаляй отдельный анализ. Верни его отдельной строкой со status=included и укажи included_in_complex.
6. ОАК/общий анализ крови без уточнений = Клинический анализ крови с лейкоцитарной формулой (5DIFF) и СОЭ (венозная кровь).
7. Коагулограмма без слова "расширенная" = Коагулограмма, скрининг. Не выбирай расширенную коагулограмму без прямого запроса.
8. Аллергены береза/ольха/орешник/лещина/плесень ищи как индивидуальные аллергены. Не выбирай аллергоПАНЕЛЬ или аллергоКОМПЛЕКС, если пациент прямо не просил панель/комплекс.
9. Орешник = лещина. Капрограмма = копрограмма. Плесень = плесневый гриб, чаще всего Alternaria alternata, а не Candida.
10. Ковид, COVID, COVID-19, коронавирус, SARS-CoV-2, мазок на ковид = лабораторный анализ на SARS-CoV-2/COVID из каталога. Если в каталоге есть ПЦР/мазок/РНК/ДНК SARS-CoV-2, выбирай его.
11. Не добавляй анализы, которых пациент не просил.
11. Если подходящей позиции нет, верни status=not_found.

Верни строго JSON:
{
  "items": [
    {
      "requested": "как понял запрос",
      "status": "found|included|not_found",
      "code": "код из каталога или пусто",
      "included_in_complex": "название комплекса, если status=included, иначе пусто",
      "comment": "коротко"
    }
  ]
}
"""

def _norm_for_self_include(text: str) -> str:
    value = normalize_text(text)
    value = re.sub(r"\b(анализ|комплекс|исследование)\b", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def _is_self_included(requested: str, included_in_complex: str, code: str, by_code: dict[str, LabTest]) -> bool:
    """
    Защита от ошибки GPT вида:
    "ПЦР-12 — входит в комплекс «ПЦР-12»".

    Если requested фактически совпадает с названием комплекса или кодом комплекса,
    это не included, а found.
    """
    req = _norm_for_self_include(requested)
    inc = _norm_for_self_include(included_in_complex)

    if req and inc and (req == inc or req in inc or inc in req):
        return True

    item = by_code.get(code)
    if item:
        item_name = _norm_for_self_include(item.name)
        if item_name and req and (req == item_name or req in item_name or item_name in req):
            return True
        if item.result_type == "комплекс" and item_name and inc and (inc == item_name or inc in item_name or item_name in inc):
            return True

    return False


def _is_covid_query_text(text: str) -> bool:
    q = normalize_text(text)
    return any(token in q for token in [
        "ковид", "covid", "коронавирус", "sars cov 2", "sars cov", "мазок на ковид", "пцр на ковид"
    ])



async def match_user_request_with_ai(raw_text: str, price_list: list[LabTest], initial_usage: AIUsage | None = None, source: str = "text") -> AIMatchResult:
    usage = AIUsage(used=False, source="")
    if initial_usage:
        usage.add(initial_usage)

    cleaned_items = clean_recognized_text_locally(raw_text)

    if not is_ai_enabled():
        # Fallback без GPT: старая локальная логика.
        fallback_items = cleaned_items or split_user_tests(raw_text)
        return AIMatchResult(
            matched_items=match_tests(fallback_items, price_list),
            cleaned_items=fallback_items,
            raw_text=raw_text,
            usage=usage,
            used_gpt_matching=False,
        )

    catalog = _build_candidate_catalog(raw_text, cleaned_items, price_list)
    if not catalog["individual_analyses"] and not catalog["complexes"]:
        fallback_items = cleaned_items or split_user_tests(raw_text)
        return AIMatchResult(
            matched_items=match_tests(fallback_items, price_list),
            cleaned_items=fallback_items,
            raw_text=raw_text,
            usage=usage,
            used_gpt_matching=False,
        )

    user_payload = {
        "raw_text": raw_text,
        "cleaned_items_local": cleaned_items,
        "catalog": catalog,
    }

    client = _client()
    response = client.chat.completions.create(
        model=OPENAI_TEXT_MODEL,
        temperature=0,
        max_tokens=1200,
        response_format={"type": "json_object"},
        messages=[
            {"role": "system", "content": GPT_MATCHING_SYSTEM_PROMPT.strip()},
            {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
        ],
    )
    usage.add(_usage_from_chat_response(response, f"{source}_gpt_match"))

    content = response.choices[0].message.content or "{}"
    try:
        data = json.loads(content)
    except Exception:
        data = {"items": []}

    lookup_price_list = list(price_list)
    existing_codes = {item.code for item in lookup_price_list if item.code}
    for complex_item in _read_complex_rows_from_xlsx():
        if complex_item.code and complex_item.code not in existing_codes:
            lookup_price_list.append(complex_item)
            existing_codes.add(complex_item.code)

    by_code = {item.code: item for item in lookup_price_list if item.code}
    matched_items: list[MatchedItem] = []

    for entry in data.get("items", []):
        requested = str(entry.get("requested") or "").strip() or "анализ"
        status = str(entry.get("status") or "").strip()
        code = str(entry.get("code") or "").strip()
        included_in_complex = str(entry.get("included_in_complex") or "").strip()

        if status == "included":
            # Анти-самоссылка: "ПЦР-12 входит в комплекс ПЦР-12" => это found-комплекс.
            if code in by_code and _is_self_included(requested, included_in_complex, code, by_code):
                candidates = find_candidate_tests(requested, lookup_price_list)
                selected = labtest_to_candidate(by_code[code], 100)
                candidates = [selected] + [c for c in candidates if c.code != code]
                matched_items.append(make_matched_from_code(requested, code, lookup_price_list, candidates=candidates))
                continue

            matched_items.append(make_included_item(requested, included_in_complex, lookup_price_list, code=code))
            continue

        if status == "found" and code in by_code:
            candidates = find_candidate_tests(requested, lookup_price_list)
            # выбранный GPT вариант ставим первым в candidates
            selected = labtest_to_candidate(by_code[code], 100)
            candidates = [selected] + [c for c in candidates if c.code != code]
            matched_items.append(make_matched_from_code(requested, code, lookup_price_list, candidates=candidates))
            continue

        matched_items.append(MatchedItem(input_name=requested, status="not_found", confidence=0))

    if not matched_items:
        fallback_items = cleaned_items or split_user_tests(raw_text)
        matched_items = match_tests(fallback_items, lookup_price_list)

    if cleaned_items:
        remember_cleaned_text(raw_text, cleaned_items)

    return AIMatchResult(
        matched_items=matched_items,
        cleaned_items=cleaned_items,
        raw_text=raw_text,
        usage=usage,
        used_gpt_matching=True,
    )
