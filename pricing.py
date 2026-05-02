import re
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from openpyxl import load_workbook
from rapidfuzz import fuzz


PRIORITY_SYNONYMS = {
    "оак": ["клинический анализ крови", "общий анализ крови"],
    "оам": ["общий анализ мочи"],
    "д димер": ["д димер", "д-димер"],
}

AUTO_LEARN_MIN_PRIORITY = 10
AUTO_LEARN_MAX_PRIORITY = 90
MANUAL_FIXED_PRIORITY = 100


@dataclass
class LabTest:
    code: str
    name: str
    result_type: str
    max_discount: int
    term_days: str
    price: int
    synonyms: list[str]
    short_name: str = ""
    group: str = ""
    priority: int = 0
    group2: str = ""
    priority2: int = 0
    auto_select_rule: str = ""
    components: str = ""


@dataclass
class CandidateItem:
    matched_name: str
    code: str
    result_type: str
    term_days: str
    price: int
    max_discount: int
    confidence: int
    short_name: str = ""
    group: str = ""
    priority: int = 0
    group2: str = ""
    priority2: int = 0
    active_group: str = ""
    active_priority_field: str = "priority"
    active_priority: int = 0
    auto_select_rule: str = ""
    components: str = ""

    @property
    def button_text(self) -> str:
        return self.short_name or self.matched_name


@dataclass
class MatchedItem:
    input_name: str
    status: str  # found | included | not_found
    matched_name: str | None = None
    code: str | None = None
    result_type: str | None = None
    term_days: str | None = None
    price: int | None = None
    max_discount: int | None = None
    confidence: int = 0
    short_name: str | None = None
    group: str | None = None
    priority: int = 0
    group2: str | None = None
    priority2: int = 0
    active_group: str = ""
    active_priority_field: str = "priority"
    active_priority: int = 0
    candidates: list[CandidateItem] = field(default_factory=list)
    auto_select_rule: str = ""
    components: str = ""
    included_in_complex: str = ""

    @property
    def button_text(self) -> str:
        return self.short_name or self.matched_name or self.input_name


def normalize_text(text: str) -> str:
    text = str(text or "").lower().replace("ё", "е")
    text = re.sub(r"[\s\u00a0]+", " ", text)
    text = re.sub(r"[.,;:()\[\]{}\"'«»]", " ", text)
    text = text.replace("-", " ").replace("—", " ")
    return re.sub(r"\s+", " ", text).strip()


def parse_int(value, default: int = 0) -> int:
    if value is None:
        return default
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return default
    try:
        number = float(text)
    except ValueError:
        digits = re.sub(r"[^0-9.]", "", text)
        if not digits:
            return default
        number = float(digits)
    return int(round(number))


def parse_discount(value) -> int:
    if value is None:
        return 0
    text = str(value).strip().replace("%", "").replace(",", ".")
    if not text:
        return 0
    try:
        number = float(text)
    except ValueError:
        return 0
    if 0 < number < 1:
        number *= 100
    return int(round(number))


def split_list_cell(value) -> list[str]:
    return [s.strip() for s in re.split(r"[;,]", str(value or "")) if s.strip()]


@lru_cache(maxsize=1)
def load_alias_groups(path: str = "alias_groups.xlsx") -> dict[str, str]:
    file_path = Path(path)
    if not file_path.exists():
        return {}

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return {}

    headers = [str(cell or "").strip() for cell in rows[0]]
    if "alias" not in headers or "group" not in headers:
        return {}

    alias_idx = headers.index("alias")
    group_idx = headers.index("group")
    aliases: dict[str, str] = {}
    for row in rows[1:]:
        alias = normalize_text(row[alias_idx] if alias_idx < len(row) else "")
        group = str(row[group_idx] if group_idx < len(row) else "").strip()
        if alias and group:
            aliases[alias] = group
    return aliases


def _load_lab_tests_from_xlsx(path: str, *, is_complex_file: bool = False) -> list[LabTest]:
    file_path = Path(path)
    if not file_path.exists():
        if is_complex_file:
            return []
        raise FileNotFoundError(f"Не найден файл прайса: {path}")

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell or "").strip() for cell in rows[0]]
    required = {"code", "name", "result_type", "max_discount", "term_days", "price", "synonyms"}
    missing = required - set(headers)
    if missing:
        filename = "complex_price_list.xlsx" if is_complex_file else "price_list.xlsx"
        raise RuntimeError(f"В {filename} не хватает колонок: {', '.join(sorted(missing))}")

    def get_cell(row, column_name: str, default=""):
        if column_name not in headers:
            return default
        idx = headers.index(column_name)
        if idx >= len(row):
            return default
        value = row[idx]
        return default if value is None else value

    items: list[LabTest] = []
    for row in rows[1:]:
        code = str(get_cell(row, "code", "")).strip()
        name = str(get_cell(row, "name", "")).strip()
        if not code and not name:
            continue
        result_type = str(get_cell(row, "result_type", "")).strip()
        components = str(get_cell(row, "components", "")).strip()
        if components and not result_type:
            result_type = "комплекс"
        items.append(
            LabTest(
                code=code,
                name=name,
                result_type=result_type,
                max_discount=parse_discount(get_cell(row, "max_discount", "")),
                term_days=str(get_cell(row, "term_days", "")).strip(),
                price=parse_int(get_cell(row, "price", "")),
                synonyms=split_list_cell(get_cell(row, "synonyms", "")),
                short_name=str(get_cell(row, "short_name", "")).strip(),
                group=str(get_cell(row, "group", "")).strip(),
                priority=parse_int(get_cell(row, "priority", ""), default=0),
                group2=str(get_cell(row, "group2", "")).strip(),
                priority2=parse_int(get_cell(row, "priority2", ""), default=0),
                auto_select_rule=str(get_cell(row, "auto_select_rule", "")).strip(),
                components=components,
            )
        )
    return items


def load_price_list(path: str = "price_list.xlsx") -> list[LabTest]:
    return _load_lab_tests_from_xlsx(path, is_complex_file=False)


def split_user_tests(text: str) -> list[str]:
    raw_parts: list[str] = []
    for line in str(text or "").splitlines():
        line = line.strip()
        if not line:
            continue
        line = line.replace(";", ",")
        raw_parts.extend(line.split(","))
    cleaned: list[str] = []
    for part in raw_parts:
        value = re.sub(r"^\s*[0-9]+[.)\-:]*\s*", "", part).strip(" -—\t")
        if value:
            cleaned.append(value)
    return cleaned


def _active_priority_for_item(item: LabTest, active_group: str = "", active_priority_field: str = "") -> tuple[int, str]:
    if active_group:
        if item.group == active_group:
            return item.priority, "priority"
        if item.group2 == active_group:
            return item.priority2, "priority2"
    if active_priority_field == "priority2":
        return item.priority2, "priority2"
    return item.priority, "priority"


def labtest_to_candidate(item: LabTest, confidence: int, *, active_group: str = "", active_priority_field: str = "") -> CandidateItem:
    active_priority, field_name = _active_priority_for_item(item, active_group, active_priority_field)
    return CandidateItem(
        matched_name=item.name,
        code=item.code,
        result_type=item.result_type,
        term_days=item.term_days,
        price=item.price,
        max_discount=item.max_discount,
        confidence=confidence,
        short_name=item.short_name,
        group=item.group,
        priority=item.priority,
        group2=item.group2,
        priority2=item.priority2,
        active_group=active_group,
        active_priority_field=field_name,
        active_priority=active_priority,
        auto_select_rule=item.auto_select_rule,
        components=item.components,
    )


def candidate_to_matched(input_name: str, candidate: CandidateItem, candidates: list[CandidateItem] | None = None, *, included_in_complex: str = "") -> MatchedItem:
    status = "included" if included_in_complex else "found"
    return MatchedItem(
        input_name=input_name,
        status=status,
        matched_name=candidate.matched_name,
        code=candidate.code,
        result_type=candidate.result_type,
        term_days=candidate.term_days,
        price=candidate.price,
        max_discount=candidate.max_discount,
        confidence=candidate.confidence,
        short_name=candidate.short_name,
        group=candidate.group,
        priority=candidate.priority,
        group2=candidate.group2,
        priority2=candidate.priority2,
        active_group=candidate.active_group,
        active_priority_field=candidate.active_priority_field,
        active_priority=candidate.active_priority,
        candidates=candidates or [],
        auto_select_rule=candidate.auto_select_rule,
        components=candidate.components,
        included_in_complex=included_in_complex,
    )


def find_item_by_code(code: str, price_list: list[LabTest]) -> LabTest | None:
    normalized_code = str(code or "").strip().lower()
    if not normalized_code:
        return None
    for item in price_list:
        if item.code.lower() == normalized_code:
            return item
    return None


def make_matched_from_code(input_name: str, code: str, price_list: list[LabTest], *, included_in_complex: str = "", candidates: list[CandidateItem] | None = None) -> MatchedItem:
    item = find_item_by_code(code, price_list)
    if not item:
        return MatchedItem(input_name=input_name, status="not_found", confidence=0)
    candidate = labtest_to_candidate(item, 100)
    return candidate_to_matched(input_name, candidate, candidates=candidates or [candidate], included_in_complex=included_in_complex)


def make_included_item(input_name: str, included_in_complex: str, price_list: list[LabTest], code: str = "") -> MatchedItem:
    if code:
        item = find_item_by_code(code, price_list)
        if item:
            candidate = labtest_to_candidate(item, 100)
            return candidate_to_matched(input_name, candidate, candidates=[candidate], included_in_complex=included_in_complex)
    return MatchedItem(input_name=input_name, status="included", matched_name=input_name, price=None, max_discount=None, included_in_complex=included_in_complex)


def get_score_for_item(query: str, raw_query: str, item: LabTest) -> int:
    if raw_query and (raw_query == item.code.lower() or query == normalize_text(item.code)):
        return 100
    best_score = 0
    variants = [item.name] + item.synonyms + [item.code, item.short_name, item.auto_select_rule]
    for variant in variants:
        if not variant:
            continue
        normalized_variant = normalize_text(variant)
        if query == normalized_variant:
            score = 100
        elif query and query in normalized_variant:
            score = max(fuzz.WRatio(query, normalized_variant), 92)
        else:
            score = fuzz.WRatio(query, normalized_variant)
        best_score = max(best_score, int(round(score)))
    return best_score


def _resolve_alias_group(query: str, alias_groups: dict[str, str]) -> str:
    exact = alias_groups.get(query)
    if exact:
        return exact
    best_group = ""
    best_score = 0
    for alias, group in alias_groups.items():
        if not alias:
            continue
        if query == alias:
            score = 120
        elif query in alias or alias in query:
            score = 100
        else:
            score = max(fuzz.WRatio(query, alias), fuzz.token_set_ratio(query, alias))
        if score > best_score:
            best_score = int(score)
            best_group = group
    return best_group if best_score >= 78 else ""


def _group_matches(item: LabTest, group: str) -> tuple[bool, str, int]:
    if item.group == group:
        return True, "priority", item.priority
    if item.group2 == group:
        return True, "priority2", item.priority2
    return False, "", 0


def _find_by_alias_group(query: str, price_list: list[LabTest], max_candidates: int) -> list[CandidateItem]:
    alias_groups = load_alias_groups()
    group = _resolve_alias_group(query, alias_groups)
    if not group:
        return []

    group_rows: list[tuple[LabTest, str, int]] = []
    for item in price_list:
        matched, field_name, active_priority = _group_matches(item, group)
        if matched:
            group_rows.append((item, field_name, active_priority))
    if not group_rows:
        return []

    def group_sort_key(row: tuple[LabTest, str, int]):
        item, field_name, active_priority = row
        item_name = normalize_text(item.name)
        item_button = normalize_text(item.short_name)
        score = get_score_for_item(query, query, item)
        starts_with_query = 1 if (item_name.startswith(query) or item_button.startswith(query)) else 0
        exact_source = 1 if query == item_name or query == item_button else 0
        component_penalty = 1 if "аллергокомпонент" in item_name else 0
        return (-active_priority, -exact_source, -starts_with_query, -score, component_penalty, item.name)

    group_rows = sorted(group_rows, key=group_sort_key)[:max_candidates]
    candidates = []
    for item, field_name, active_priority in group_rows:
        score = max(100, get_score_for_item(query, query, item))
        candidate = labtest_to_candidate(item, score, active_group=group, active_priority_field=field_name)
        candidates.append(candidate)
    return candidates


def _rule_match_score(query: str, rule: str) -> int:
    rule = normalize_text(rule)
    if not query or not rule:
        return 0
    if query == rule:
        return 1000 + len(rule)
    query_tokens = set(query.split())
    rule_tokens = set(rule.split())
    if rule_tokens and rule_tokens.issubset(query_tokens):
        return 900 + len(rule)
    if rule in query:
        return 800 + len(rule)
    if query in rule and len(query) >= 4:
        return 500 + len(query)
    return 0


def smart_default_select(query: str, candidates: list[CandidateItem]) -> list[CandidateItem]:
    if not candidates:
        return candidates
    scored_candidates: list[tuple[int, CandidateItem]] = []
    for candidate in candidates:
        rule_scores = [_rule_match_score(query, rule) for rule in split_list_cell(candidate.auto_select_rule)]
        rule_score = max(rule_scores) if rule_scores else 0
        short_name_score = _rule_match_score(query, candidate.short_name)
        name_score = _rule_match_score(query, candidate.matched_name)
        total_score = max(rule_score, short_name_score, name_score)
        scored_candidates.append((total_score, candidate))
    best_score = max(score for score, _ in scored_candidates)
    if best_score <= 0:
        return candidates
    best_candidate = sorted(
        [candidate for score, candidate in scored_candidates if score == best_score],
        key=lambda candidate: (-candidate.active_priority, -candidate.priority, candidate.matched_name),
    )[0]
    return [best_candidate] + [candidate for candidate in candidates if candidate.code != best_candidate.code]


def find_candidate_tests(input_name: str, price_list: list[LabTest], min_score: int = 80, close_gap: int = 8, max_candidates: int = 8) -> list[CandidateItem]:
    query = normalize_text(input_name)
    raw_query = str(input_name or "").strip().lower()
    if not query and not raw_query:
        return []

    group_candidates = _find_by_alias_group(query, price_list, max_candidates=max_candidates)
    if group_candidates:
        return smart_default_select(query, group_candidates)

    priority_keywords = PRIORITY_SYNONYMS.get(query)
    working_price_list = price_list
    if priority_keywords:
        filtered_price_list: list[LabTest] = []
        for item in price_list:
            combined_text = normalize_text(item.name + " " + " ".join(item.synonyms))
            if any(keyword in combined_text for keyword in priority_keywords):
                filtered_price_list.append(item)
        if filtered_price_list:
            working_price_list = filtered_price_list

    scored: list[tuple[int, LabTest]] = []
    for item in working_price_list:
        score = get_score_for_item(query, raw_query, item)
        if score >= min_score:
            scored.append((score, item))
    if not scored:
        return []

    scored.sort(key=lambda pair: (-pair[0], -pair[1].priority, pair[1].name))
    best_score = scored[0][0]
    close_items = [(score, item) for score, item in scored if best_score - score <= close_gap][:max_candidates]
    candidates = [labtest_to_candidate(item, score) for score, item in close_items]
    return smart_default_select(query, candidates)


def match_one_test(input_name: str, price_list: list[LabTest]) -> MatchedItem:
    candidates = find_candidate_tests(input_name, price_list)
    if not candidates:
        query = normalize_text(input_name)
        raw_query = str(input_name or "").strip().lower()
        best_score = 0
        for item in price_list:
            best_score = max(best_score, get_score_for_item(query, raw_query, item))
        return MatchedItem(input_name=input_name, status="not_found", confidence=best_score)
    return candidate_to_matched(input_name, candidates[0], candidates=candidates)


def match_tests(user_tests: list[str], price_list: list[LabTest]) -> list[MatchedItem]:
    return [match_one_test(test, price_list) for test in user_tests]


def has_ambiguous_items(items: list[MatchedItem]) -> bool:
    return False


def calculate_item_price(item: MatchedItem, patient_discount: int) -> dict:
    if item.status == "included":
        return {
            "input_name": item.input_name,
            "status": "included",
            "code": item.code,
            "name": item.matched_name or item.input_name,
            "included_in_complex": item.included_in_complex,
            "final_price": None,
            "discount_used": None,
        }
    if item.status != "found" or item.price is None or item.max_discount is None:
        return {"input_name": item.input_name, "status": item.status, "final_price": None, "discount_used": None}
    discount_used = min(patient_discount, item.max_discount)
    final_price = round(item.price * (100 - discount_used) / 100)
    return {
        "input_name": item.input_name,
        "status": "found",
        "code": item.code,
        "name": item.matched_name,
        "short_name": item.short_name or "",
        "result_type": item.result_type or "не указано",
        "term_days": item.term_days or "не указано",
        "base_price": item.price,
        "max_discount": item.max_discount,
        "discount_used": discount_used,
        "final_price": final_price,
    }


def learn_group_priority(active_group: str, selected_code: str, price_list_path: str = "price_list.xlsx") -> bool:
    """
    Самообучение ранжирования внутри активной группы.

    Правила:
    - приоритет 100 не меняется автоматически;
    - выбранный анализ с приоритетом <100 получает +1, но не выше 90;
    - остальные анализы активной группы с приоритетом <100 получают -1, но не ниже 10;
    - если группа совпала через group, меняем priority;
    - если группа совпала через group2, меняем priority2.
    """
    active_group = str(active_group or "").strip()
    selected_code = str(selected_code or "").strip()
    if not active_group or not selected_code:
        return False

    path = Path(price_list_path)
    if not path.exists():
        return False

    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=False))
    if not rows:
        return False

    headers = [str(cell.value or "").strip() for cell in rows[0]]
    required = ["code", "group", "priority", "group2", "priority2"]
    for col in required:
        if col not in headers:
            return False

    idx = {name: headers.index(name) for name in headers}
    changed = False

    target_rows: list[tuple[int, str, int]] = []
    for row_index, row in enumerate(rows[1:], start=2):
        code = str(row[idx["code"]].value or "").strip()
        group = str(row[idx["group"]].value or "").strip()
        group2 = str(row[idx["group2"]].value or "").strip()
        if group == active_group:
            target_rows.append((row_index, "priority", parse_int(row[idx["priority"]].value, 0)))
        if group2 == active_group:
            target_rows.append((row_index, "priority2", parse_int(row[idx["priority2"]].value, 0)))

    if not target_rows:
        return False

    for row_index, field_name, old_value in target_rows:
        row = rows[row_index - 1]
        code = str(row[idx["code"]].value or "").strip()
        if old_value >= MANUAL_FIXED_PRIORITY:
            continue
        if code == selected_code:
            new_value = min(AUTO_LEARN_MAX_PRIORITY, old_value + 1)
        else:
            new_value = max(AUTO_LEARN_MIN_PRIORITY, old_value - 1)
        if new_value != old_value:
            row[idx[field_name]].value = new_value
            changed = True

    if changed:
        workbook.save(path)
        load_price_list.cache_clear() if hasattr(load_price_list, "cache_clear") else None
        load_alias_groups.cache_clear()
    return changed
