
import re
from pathlib import Path
from typing import Iterable, Any

from openai import OpenAI
from openpyxl import load_workbook
from rapidfuzz import fuzz

from config import OPENAI_API_KEY, OPENAI_TEXT_MODEL

client = OpenAI(api_key=OPENAI_API_KEY)
MODEL = OPENAI_TEXT_MODEL or "gpt-4.1-nano"


SYSTEM_PROMPT = """
Ты — опытный медицинский ассистент лаборатории.

Твоя задача: из текста пациента составить корректный список лабораторных анализов и комплексов СТРОГО по переданному каталогу.

Каталог состоит из:
1) ОДИНОЧНЫХ АНАЛИЗОВ.
2) КОМПЛЕКСОВ, где после названия комплекса указан его состав.

ВАЖНЫЕ ПРАВИЛА:

1. Удаляй мусор и не-лабораторные исследования:
консультации, приемы врачей, УЗИ, ЭКГ, рентген, МРТ, КТ, ФГДС, колоноскопия, процедуры, вводные слова.

2. Если это похоже на лабораторный анализ, но его нет в каталоге:
верни строку:
<исходный запрос> — требует уточнения у администратора

3. Не дублируй.
Один и тот же анализ или комплекс должен быть в списке только один раз.

4. Комплексы.
Если пациент запросил несколько анализов, которые входят в один комплекс из каталога, выбери комплекс.
Пример:
ПТ + МНО + АЧТВ + ТВ + ФГ => Коагулограмма, скрининг.

Если пациент запросил комплекс и отдельно анализ, который входит в этот комплекс, НЕ добавляй отдельный анализ как платную позицию.
Верни его только так:
<анализ> — входит в комплекс "<точное название комплекса>"

5. Дефолты:
- ОАК, общий анализ крови, клинический анализ крови => Клинический анализ крови с лейкоцитарной формулой (5DIFF) и СОЭ (венозная кровь)
- ребенок / ребёнок / из пальца / капиллярная кровь => капиллярный ОАК с лейкоформулой и СОЭ.
- Коагулограмма без слова "расширенная" => Коагулограмма, скрининг.
- Не выбирай расширенную коагулограмму без прямого запроса "расширенная".

6. Аллергены:
Если пациент написал "береза", "ольха", "орешник", "лещина", "плесень" и т.п., ищи именно индивидуальные аллергены из каталога.
Не заменяй индивидуальный аллерген на панель/комплекс, если пациент прямо не просил панель.

7. Нельзя:
- придумывать анализы,
- менять названия из каталога,
- писать объяснения,
- писать reasoning,
- делать самоссылки вида "ОАК входит в комплекс ОАК".

8. Формат ответа:
Верни ТОЛЬКО список, каждая позиция с новой строки.
Без нумерации.
Без пояснений.

Допустимые форматы строк:
<точное название из каталога>
<исходный анализ> — входит в комплекс "<точное название комплекса из каталога>"
<исходный анализ> — требует уточнения у администратора
"""


COAG_TOKENS = {"коагулограмма", "коагулограмму", "пт", "мно", "ачтв", "тв", "фг", "фибриноген", "протромбин", "тромбиновое"}
CBC_TOKENS = {"оак", "общий анализ крови", "клинический анализ крови", "анализ крови"}
ALLERGEN_STEMS = {
    "берез": ["берез", "берёз", "betula"],
    "ольх": ["ольх", "alnus"],
    "ореш": ["ореш", "лещин", "corylus", "hazel"],
    "лещин": ["ореш", "лещин", "corylus", "hazel"],
    "плес": ["плес", "плеснев", "гриб", "alternaria", "cladosporium", "aspergillus", "penicillium"],
}


def _usage_total(usage: Any) -> int:
    if usage is None:
        return 0
    value = getattr(usage, "total_tokens", None)
    if value is not None:
        return int(value)
    try:
        return int(usage.get("total_tokens", 0))
    except Exception:
        return 0


def normalize_text(text: str) -> str:
    text = str(text or "").lower().replace("ё", "е")
    text = re.sub(r"[.,;:()\[\]{}\"'«»/\\\\]", " ", text)
    text = text.replace("-", " ").replace("—", " ")
    return re.sub(r"\s+", " ", text).strip()


def _tokens(text: str) -> set[str]:
    return {t for t in normalize_text(text).split() if len(t) >= 2}


def load_complex_catalog(path: str = "complex_price_list.xlsx") -> list[dict]:
    file_path = Path(path)
    if not file_path.exists():
        return []

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(c or "").strip() for c in rows[0]]

    def idx(name: str):
        return headers.index(name) if name in headers else None

    name_i = idx("name")
    comp_i = idx("components")
    price_i = idx("price")
    term_i = idx("term_days")

    if name_i is None:
        return []

    result = []
    for row in rows[1:]:
        name = row[name_i] if name_i < len(row) else ""
        if not name:
            continue
        result.append({
            "name": str(name).strip(),
            "components": str(row[comp_i] if comp_i is not None and comp_i < len(row) and row[comp_i] else "").strip(),
            "price": str(row[price_i] if price_i is not None and price_i < len(row) and row[price_i] else "").strip(),
            "term_days": str(row[term_i] if term_i is not None and term_i < len(row) and row[term_i] else "").strip(),
        })
    return result


def _is_panel_or_complex_name(name: str) -> bool:
    low = normalize_text(name)
    return any(x in low for x in ["панель", "комплекс", "immunocap"])


def _allergen_query_stems(query: str) -> list[str]:
    q = normalize_text(query)
    found = []
    for stem, variants in ALLERGEN_STEMS.items():
        if any(v in q for v in variants):
            found.append(stem)
    return found


def _item_matches_allergen(item_name: str, stems: list[str]) -> bool:
    n = normalize_text(item_name)
    for stem in stems:
        variants = ALLERGEN_STEMS.get(stem, [stem])
        if any(v in n for v in variants):
            return True
    return False


def _score_item_for_query(query: str, item) -> int:
    name = str(getattr(item, "name", "") or "")
    short_name = str(getattr(item, "short_name", "") or "")
    text = f"{name} {short_name}"
    q = normalize_text(query)
    t = normalize_text(text)
    if not q or not t:
        return 0
    if q in t:
        return 100
    return int(fuzz.WRatio(q, t))


def build_catalog_for_gpt(price_list, user_text: str, max_single: int = 220, max_complex: int = 80) -> str:
    """
    ВАЖНО:
    Не отправляем GPT первые 900 строк прайса подряд.
    Собираем релевантный каталог под конкретный запрос:
    - ОАК-дефолты;
    - коагулограммы и комплексы по составу;
    - индивидуальные аллергены по нужным словам;
    - fuzzy-кандидаты по названию.
    """
    q = normalize_text(user_text)
    q_tokens = _tokens(user_text)

    selected_items = []
    seen_names = set()

    def add_item(item):
        name = str(getattr(item, "name", "") or "").strip()
        if not name:
            return
        key = normalize_text(name)
        if key not in seen_names:
            seen_names.add(key)
            selected_items.append(item)

    # 1. ОАК: обязательно добавляем правильные варианты, чтобы GPT не выбрал случайные CBC.
    if "оак" in q_tokens or "общий анализ крови" in q or "клинический анализ крови" in q or "анализ крови" in q:
        for item in price_list:
            n = normalize_text(getattr(item, "name", ""))
            if "клинический анализ крови" in n and "соэ" in n:
                add_item(item)

    # 2. Коагулограмма/компоненты.
    if q_tokens & COAG_TOKENS:
        for item in price_list:
            n = normalize_text(getattr(item, "name", ""))
            if any(tok in n for tok in ["мно", "ачтв", "фибриноген", "протромбин", "тромбиновое"]):
                add_item(item)

    # 3. Аллергены: индивидуальные позиции, не панели/комплексы.
    allergen_stems = _allergen_query_stems(user_text)
    if allergen_stems:
        for item in price_list:
            name = str(getattr(item, "name", "") or "")
            if _item_matches_allergen(name, allergen_stems) and not _is_panel_or_complex_name(name):
                add_item(item)

    # 4. Fuzzy-кандидаты по всему прайсу.
    scored = []
    for item in price_list:
        score = _score_item_for_query(user_text, item)
        if score >= 58:
            scored.append((score, item))
    scored.sort(key=lambda x: (-x[0], str(getattr(x[1], "name", ""))))
    for _, item in scored[:max_single]:
        add_item(item)

    # Ограничение одиночных.
    selected_items = selected_items[:max_single]

    # Комплексы: отбираем по названию и составу, плюс обязательно коагулограмму при наличии ее компонентов.
    complexes = load_complex_catalog()
    selected_complexes = []
    seen_complex = set()

    def add_complex(c):
        key = normalize_text(c["name"])
        if key not in seen_complex:
            seen_complex.add(key)
            selected_complexes.append(c)

    for c in complexes:
        blob = normalize_text(c["name"] + " " + c.get("components", ""))
        score = fuzz.WRatio(q, blob) if q else 0

        if q_tokens & COAG_TOKENS and "коагулограмма" in normalize_text(c["name"]):
            add_complex(c)
            continue

        if any(tok in blob for tok in q_tokens if len(tok) >= 3):
            add_complex(c)
            continue

        if score >= 62:
            add_complex(c)

    selected_complexes = selected_complexes[:max_complex]

    lines: list[str] = []
    lines.append("ОДИНОЧНЫЕ АНАЛИЗЫ:")
    for item in selected_items:
        name = getattr(item, "name", "")
        price = getattr(item, "price", "")
        term = getattr(item, "term_days", "")
        if name:
            lines.append(f"- {name} | цена: {price} | срок: {term}")

    if selected_complexes:
        lines.append("")
        lines.append("КОМПЛЕКСЫ:")
        for item in selected_complexes:
            line = f"- {item['name']} | цена: {item['price']} | срок: {item['term_days']}"
            if item["components"]:
                line += f" | состав: {item['components']}"
            lines.append(line)

    return "\n".join(lines)


def call_gpt(user_text: str, catalog: str):
    response = client.responses.create(
        model=MODEL,
        temperature=0,
        max_output_tokens=500,
        input=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": f"Запрос пациента:\n{user_text}\n\nКаталог:\n{catalog}"},
        ],
    )
    return response.output_text, response.usage


def clean_gpt_output(text: str) -> list[str]:
    lines = [str(line).strip() for line in str(text or "").splitlines()]
    cleaned = []
    seen = set()

    for line in lines:
        if not line:
            continue
        line = re.sub(r"^\s*\d+[\).\-\s]+", "", line).strip()
        if not line:
            continue

        if "входит в комплекс" in line.lower():
            left = normalize_text(line.split("—")[0] if "—" in line else line.split("-")[0])
            right = normalize_text(line.split("комплекс", 1)[-1])
            if left and left in right:
                continue

        key = normalize_text(line)
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(line)

    return cleaned


def enforce_business_rules(lines: list[str]) -> list[str]:
    result = []
    has_coag_screening = any("коагулограмма, скрининг" in line.lower() for line in lines)

    for line in lines:
        low = line.lower()

        # Без явного запроса расширенной коагулограммы GPT иногда выбирает расширенную.
        if "коагулограмма" in low and "расширенная" in low:
            line = re.sub("Коагулограмма, расширенная", "Коагулограмма, скрининг", line, flags=re.IGNORECASE)

        if has_coag_screening and any(x in low for x in ["мно", "ачтв", "фибриноген", "тромбиновое", "протромбин"]):
            if "входит в комплекс" not in low and "коагулограмма" not in low:
                line = f'{line} — входит в комплекс "Коагулограмма, скрининг"'

        result.append(line)

    deduped = []
    seen = set()
    for line in result:
        key = normalize_text(line)
        if key not in seen:
            seen.add(key)
            deduped.append(line)
    return deduped


def validate_against_catalog(lines: list[str], catalog_names: Iterable[str]) -> list[str]:
    catalog_names = [str(x) for x in catalog_names if str(x).strip()]
    catalog_norm = {normalize_text(x): x for x in catalog_names}

    complex_names = [x["name"] for x in load_complex_catalog()]
    complex_norm = {normalize_text(x): x for x in complex_names}

    validated = []

    for line in lines:
        low = line.lower()

        if "входит в комплекс" in low or "требует уточнения" in low:
            validated.append(line)
            continue

        norm = normalize_text(line)

        if norm in catalog_norm:
            validated.append(catalog_norm[norm])
            continue

        if norm in complex_norm:
            validated.append(complex_norm[norm])
            continue

        found = None
        for n_norm, original in {**catalog_norm, **complex_norm}.items():
            if n_norm and n_norm in norm:
                found = original
                break

        if found:
            validated.append(found)
        else:
            validated.append(f"{line} — требует уточнения у администратора")

    return validated


def gpt_match(user_text: str, price_list, catalog_names: list[str]):
    catalog = build_catalog_for_gpt(price_list, user_text)
    raw, usage = call_gpt(user_text, catalog)

    cleaned = clean_gpt_output(raw)
    business_fixed = enforce_business_rules(cleaned)

    all_catalog_names = list(catalog_names) + [x["name"] for x in load_complex_catalog()]
    validated = validate_against_catalog(business_fixed, all_catalog_names)

    return validated, usage, _usage_total(usage)
